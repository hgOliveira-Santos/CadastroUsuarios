import customtkinter as ctk  # Importação da biblioteca customizada customtkinter
from tkinter import messagebox  # Importação de messagebox do módulo tkinter
import openpyxl  # Importação do módulo openpyxl para manipulação de planilhas Excel
import pathlib  # Importação de pathlib para manipulação de caminhos de arquivos
from openpyxl import Workbook  # Importação da classe Workbook do módulo openpyxl

# Configurações iniciais da aparência usando a biblioteca customizada
ctk.set_appearance_mode("System")
ctk.set_default_color_theme("blue")

# Classe principal da aplicação que herda de ctk.CTk
class App(ctk.CTk):
    def __init__(self):
        super().__init__()  # Inicializa a classe pai (ctk.CTk)
        self.layout_config()  # Configuração inicial da interface
        self.aparência()  # Configuração da aparência da interface
        self.todo_sistema()  # Configuração dos elementos principais da aplicação

    # Configuração geral da interface
    def layout_config(self):
        self.title("Sistema de Cadastro de Clientes")  # Define o título da janela
        self.geometry("700x500")  # Define o tamanho inicial da janela

    # Configuração dos elementos visuais da aparência
    def aparência(self):
        # Label e OptionMenu para escolher o tema da aplicação
        self.lb_aparência = ctk.CTkLabel(master=self, text="Tema", font=("Century Gothic bold", 14), bg_color="transparent", text_color=["#000", "#fff"]).place(x=50, y=430)
        self.opção_tema = ctk.CTkOptionMenu(master=self, values=["Light", "Dark", "System"], command=self.troca_aparência).place(x=50, y=460)

    # Configuração dos elementos principais da aplicação
    def todo_sistema(self):
        # Cabeçalho da aplicação
        frame = ctk.CTkFrame(master=self, width=700, height=50, corner_radius=0, bg_color="teal", fg_color="teal").place(x=0, y=10)
        título = ctk.CTkLabel(master=frame, text="Sistema de Cadastro de Clientes", font=("Century Gothic bold", 24), text_color="#fff", bg_color="teal").place(x=180, y=20)
        span = ctk.CTkLabel(master=self, text="Preencha todos os campos do formulário!", font=("Century Gothic bold", 16), text_color=["#000", "#fff"]).place(x=50, y=70)

        # Variáveis de texto
        self.nome_value = ctk.StringVar()
        self.contato_value = ctk.StringVar()
        self.endereço_value = ctk.StringVar()
        self.idade_value = ctk.StringVar()

        # Entradas de texto para os dados do cliente
        nome_entry = ctk.CTkEntry(master=self, width=250, textvariable=self.nome_value, font=("Century Gothic", 16), fg_color="transparent")
        nome_entry.place(x=50, y=160)
        
        contato_entry = ctk.CTkEntry(master=self, width=200, textvariable=self.contato_value, font=("Century Gothic", 16), fg_color="transparent")
        contato_entry.place(x=350, y=160)
        
        endereco_entry = ctk.CTkEntry(master=self, width=250, textvariable=self.endereço_value, font=("Century Gothic", 16), fg_color="transparent")
        endereco_entry.place(x=50, y=240)
        
        idade_entry = ctk.CTkEntry(master=self, width=100, textvariable=self.idade_value, font=("Century Gothic", 16), fg_color="transparent")
        idade_entry.place(x=350, y=240)

        # Labels explicativos para cada campo de entrada
        lb_nome = ctk.CTkLabel(master=self, text="Nome completo:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_nome.place(x=50, y=120)
        
        lb_contato = ctk.CTkLabel(master=self, text="Contato:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_contato.place(x=350, y=120)
        
        lb_endereco = ctk.CTkLabel(master=self, text="Endereço:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_endereco.place(x=50, y=200)
        
        lb_idade = ctk.CTkLabel(master=self, text="Idade:", font=("Century Gothic bold", 16), text_color=["#000", "#fff"])
        lb_idade.place(x=350, y=200)

        # Botões para salvar e limpar dados
        botao_salvar = ctk.CTkButton(master=self, text="Salvar dados".upper(), command=self.salvar_dados, fg_color="#006464", hover_color="#003232")
        botao_salvar.place(x=360, y=460)

        botao_limpar_dados = ctk.CTkButton(master=self, text="Limpar".upper(), command=self.limpar_dados, fg_color="#555", hover_color="#333")
        botao_limpar_dados.place(x=510, y=460)

    # Método para criar o arquivo Excel se não existir
    def define_arquivo(self):
        caminho_arquivo = pathlib.Path("Clientes.xlsx")
        if not caminho_arquivo.exists():
            planilha = Workbook()
            planilha_ativa = planilha.active

            # Define os cabeçalhos da planilha
            planilha_ativa["A1"] = "Nome completo"
            planilha_ativa["B1"] = "Contato"
            planilha_ativa["C1"] = "Endereço"
            planilha_ativa["D1"] = "Idade"
            planilha.save("Clientes.xlsx")

    # Método para salvar os dados preenchidos na planilha Excel
    def salvar_dados(self):
        nome = self.nome_value.get()
        contato = self.contato_value.get()
        endereço = self.endereço_value.get()
        idade = self.idade_value.get()

        # Verifica se todos os campos foram preenchidos
        if nome == "" or contato == "" or endereço == "" or idade == "":
            messagebox.showerror("Sistema", "Erro!\nPor favor preencha todos os campos.")
        else:
            try:
                # Carrega a planilha existente
                planilha = openpyxl.load_workbook("Clientes.xlsx")
                planilha_ativa = planilha.active

                # Insere os dados na próxima linha disponível
                planilha_ativa.cell(column=1, row=planilha_ativa.max_row+1, value=nome)
                planilha_ativa.cell(column=2, row=planilha_ativa.max_row, value=contato)        
                planilha_ativa.cell(column=3, row=planilha_ativa.max_row, value=endereço)
                planilha_ativa.cell(column=4, row=planilha_ativa.max_row, value=idade)

                planilha.save(r"Clientes.xlsx")  # Salva a planilha atualizada
                messagebox.showinfo("Sistema", "Dados salvos com sucesso!")  # Exibe mensagem de sucesso
                self.destroy()  # Fecha a janela após salvar
    
            except Exception:
                messagebox.showerror("Erro", "Ocorreu um erro ao salvar os dados")  # Exibe mensagem de erro em caso de falha

    # Método para limpar todos os campos de entrada
    def limpar_dados(self):
        self.nome_value.set("")
        self.contato_value.set("")
        self.endereço_value.set("")
        self.idade_value.set("")

    # Método para trocar a aparência da aplicação conforme escolha do usuário
    def troca_aparência(self, nova_aparencia):
        ctk.set_appearance_mode(nova_aparencia)

# Ponto de entrada da aplicação
if __name__ == "__main__":
    app = App()  # Cria uma instância da classe principal App
    app.define_arquivo()  # Verifica e cria o arquivo Excel se não existir
    app.mainloop()  # Inicia o loop principal da interface gráfica